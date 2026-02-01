"""
Office document embedded label writer.

Supports DOCX, XLSX, PPTX (all are ZIP-based OOXML formats).
Stores labels in Custom Document Properties with name "openlabels".
"""

from pathlib import Path
from typing import Optional

from ...core.labels import LabelSet
from .base import EmbeddedLabelWriter, logger


class OfficeLabelWriter(EmbeddedLabelWriter):
    """
    Write/read labels to Office document custom properties.

    Supports DOCX, XLSX, PPTX (all are ZIP-based OOXML formats).
    Stores in Custom Document Properties with name "openlabels".
    """

    SUPPORTED_EXTENSIONS = {'.docx', '.xlsx', '.pptx'}

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def write(self, path: Path, label_set: LabelSet) -> bool:
        """Write LabelSet to Office custom properties."""
        suffix = path.suffix.lower()

        try:
            if suffix == '.docx':
                return self._write_docx(path, label_set)
            elif suffix == '.xlsx':
                return self._write_xlsx(path, label_set)
            elif suffix == '.pptx':
                return self._write_pptx(path, label_set)
            return False
        except (OSError, ValueError, KeyError) as e:
            logger.error(f"Failed to write Office labels: {e}")
            return False

    def read(self, path: Path) -> Optional[LabelSet]:
        """Read LabelSet from Office custom properties."""
        suffix = path.suffix.lower()

        try:
            if suffix == '.docx':
                return self._read_docx(path)
            elif suffix == '.xlsx':
                return self._read_xlsx(path)
            elif suffix == '.pptx':
                return self._read_pptx(path)
            return None
        except (OSError, ValueError, KeyError) as e:
            logger.debug(f"No labels found in Office doc: {e}")
            return None

    def _write_docx(self, path: Path, label_set: LabelSet) -> bool:
        """Write to DOCX custom properties."""
        try:
            from docx import Document
        except ImportError:
            logger.warning("python-docx not installed, cannot write DOCX labels")
            return False

        doc = Document(path)
        core_props = doc.core_properties
        # Use custom_properties if available, otherwise fall back to comments
        # python-docx doesn't have great custom property support,
        # so we'll use the comments field as a workaround
        core_props.comments = label_set.to_json(compact=True)
        doc.save(path)
        return True

    def _read_docx(self, path: Path) -> Optional[LabelSet]:
        """Read from DOCX custom properties."""
        try:
            from docx import Document
        except ImportError:
            return None

        doc = Document(path)
        comments = doc.core_properties.comments
        if comments and comments.startswith('{"v":'):
            return LabelSet.from_json(comments)
        return None

    def _write_xlsx(self, path: Path, label_set: LabelSet) -> bool:
        """Write to XLSX custom properties."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl not installed, cannot write XLSX labels")
            return False

        wb = load_workbook(path)
        # Use custom doc properties
        if wb.properties is None:
            from openpyxl.packaging.core import DocumentProperties
            wb.properties = DocumentProperties()
        wb.properties.description = label_set.to_json(compact=True)
        wb.save(path)
        return True

    def _read_xlsx(self, path: Path) -> Optional[LabelSet]:
        """Read from XLSX custom properties."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return None

        wb = load_workbook(path, read_only=True)
        if wb.properties and wb.properties.description:
            desc = wb.properties.description
            if desc.startswith('{"v":'):
                return LabelSet.from_json(desc)
        return None

    def _write_pptx(self, path: Path, label_set: LabelSet) -> bool:
        """Write to PPTX custom properties."""
        try:
            from pptx import Presentation
        except ImportError:
            logger.warning("python-pptx not installed, cannot write PPTX labels")
            return False

        prs = Presentation(path)
        prs.core_properties.comments = label_set.to_json(compact=True)
        prs.save(path)
        return True

    def _read_pptx(self, path: Path) -> Optional[LabelSet]:
        """Read from PPTX custom properties."""
        try:
            from pptx import Presentation
        except ImportError:
            return None

        prs = Presentation(path)
        comments = prs.core_properties.comments
        if comments and comments.startswith('{"v":'):
            return LabelSet.from_json(comments)
        return None


__all__ = ['OfficeLabelWriter']
